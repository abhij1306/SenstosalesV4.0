import contextlib
import io
import logging
import os
import sys
from copy import copy

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.cell_range import CellRange

from backend.core.num_to_words import amount_to_words
from backend.services.excel_writer import ExcelWriter

logger = logging.getLogger(__name__)

class InvoiceExcel(ExcelWriter):
    def __init__(self, header, items, db, save_path):
        data = {"header": header, "items": items, "save_path": save_path}
        # We don't use xlsxwriter for Invoice, so we don't call super().__init__ which creates one
        self.data = data
        self.db = db
        self.output = io.BytesIO()

    def generate_excel(self, filename=None):
        """
        Overridden to use openpyxl and template GST_INV_31.xlsx.
        Restored from backup generate_exact_invoice_excel.
        """
        header = self.data["header"]
        items = self.data["items"]
        db = self.db
        save_path = self.data.get("save_path")

        if not filename:
            filename = f"Invoice_{header.get('invoice_number', 'Draft')}.xlsx"

        # Template resolution
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            template_path = os.path.join(sys._MEIPASS, "backend", "templates", "GST_INV_31.xlsx")
        else:
            template_path = os.path.join(os.path.dirname(__file__), "..", "templates", "GST_INV_31.xlsx")

        if not os.path.exists(template_path):
            logger.error(f"Template GST_INV_31.xlsx not found at {template_path}")
            # Fallback to base xlsxwriter if template missing? 
            # No, user wants correct format. 
            raise FileNotFoundError(f"Template not found: {template_path}")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        def set_val(coord, value, bold=False, align=None):
            try:
                cell = ws[coord]
                if isinstance(cell, MergedCell):
                    for m_range in ws.merged_cells.ranges:
                        if coord in m_range:
                            cell = ws.cell(row=m_range.min_row, column=m_range.min_col)
                            break
                if value is not None:
                    cell.value = value
                if bold:
                    cell.font = Font(name="Calibri", size=10, bold=True)
                if align:
                    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            except Exception as e:
                logger.warning(f"Error setting value at {coord}: {e}")

        # 1. Fetch Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        # 2. Header Info
        set_val("L3", header.get("invoice_number", ""), align="left")
        set_val("P3", header.get("invoice_date", ""), align="left")
        
        gem_num = header.get("gemc_number", "")
        if gem_num:
            set_val("I4", f"GEMC: {gem_num}", align="left")
            set_val("O4", f"Date: {header.get('gemc_date', '')}", align="left")
        set_val("R4", header.get("payment_terms", ""), align="left")
        
        set_val("L6", str(header.get("dc_number", "") or ""), align="left")
        set_val("P6", header.get("dc_date") or header.get("invoice_date", ""), align="left")
        
        set_val("L8", str(header.get("po_numbers", "") or ""), align="left")
        set_val("P8", str(header.get("buyers_order_date", "") or ""), align="left")
        
        set_val("L10", header.get("despatch_doc_no", ""), align="left")
        set_val("O10", header.get("srv_number") or header.get("srv_no", ""), align="left")
        srv_dt = header.get("srv_date", "")
        if srv_dt and len(srv_dt) > 10: srv_dt = srv_dt[:10]
        set_val("Q10", srv_dt, align="left")
        
        set_val("L12", (header.get("despatch_through", "") or "").upper(), align="left")
        set_val("Q12", (header.get("destination", "") or "").upper(), align="left")

        # 3. Seller Info
        s_name = header.get("supplier_name") or settings.get("supplier_name", "")
        s_addr = header.get("supplier_address") or settings.get("supplier_address", "")
        s_gstin = header.get("supplier_gstin") or settings.get("supplier_gstin", "")
        s_state = settings.get("supplier_state", "")
        s_code = settings.get("supplier_state_code", "")
        s_contact = header.get("supplier_contact") or settings.get("supplier_contact", "")
        
        set_val("A3", s_name, align="left")
        set_val("A4", s_addr, align="left")
        set_val("C5", s_gstin, align="left")
        set_val("C6", s_state, align="left")
        set_val("F6", s_code, align="left")
        with contextlib.suppress(Exception):
            ws.merge_cells("C7:F7")
        set_val("C7", s_contact, align="left")

        # 4. Buyer Info
        b_name = header.get("consignee_name") or header.get("buyer_name")
        b_addr = header.get("consignee_address") or header.get("buyer_address")
        b_gst = header.get("consignee_gstin") or header.get("buyer_gstin")
        b_place = header.get("place_of_supply")
        b_state = header.get("buyer_state")

        if not b_name:
             try:
                 buyer_row = db.execute("SELECT name, address, gstin, place_of_supply, state FROM buyers WHERE is_default = 1 LIMIT 1").fetchone()
                 if buyer_row:
                     b_name, b_addr, b_gst, b_place, b_state = buyer_row["name"], buyer_row["address"], buyer_row["gstin"], buyer_row["place_of_supply"], buyer_row["state"]
             except Exception: pass

        set_val("A9", b_name or "BUYER NAME", align="left")
        set_val("A10", b_addr or "BUYER ADDRESS", align="left")
        set_val("C11", b_gst or "BUYER GSTIN", align="left")
        set_val("C12", b_state or "", align="left")
        set_val("C13", b_place or "", align="left")

        # 5. Line Items
        start_row = 17
        template_total_row = 23
        template_capacity = 6
        num_items = len(items)
        rows_to_insert = max(0, num_items - template_capacity)
        
        for r_clear in range(start_row, template_total_row):
            for c_clear in range(1, 20):
                cell = ws.cell(row=r_clear, column=c_clear)
                if not isinstance(cell, MergedCell): cell.value = None
        
        if rows_to_insert > 0:
            try:
                m_ranges = [CellRange(m.coord) for m in ws.merged_cells.ranges]
                ws.merged_cells.clear()
                ws.insert_rows(start_row + 1, amount=rows_to_insert)
                for m in m_ranges:
                    if m.min_row >= start_row + 1: m.shift(row_shift=rows_to_insert)
                    ws.merged_cells.add(m)
            except Exception: pass
        elif num_items < template_capacity:
             for r_hide in range(start_row + num_items, template_total_row):
                 ws.row_dimensions[r_hide].hidden = True

        t_qty, t_taxable, t_cgst, t_sgst, t_total = 0, 0, 0, 0, 0
        cgst_rate = float(settings.get("cgst_rate", 9.0))
        sgst_rate = float(settings.get("sgst_rate", 9.0))

        for idx, item in enumerate(items):
            r = start_row + idx
            if r > start_row and rows_to_insert > 0:
                 for col in range(1, 26):
                    src_c = ws.cell(row=start_row, column=col)
                    dest_c = ws.cell(row=r, column=col)
                    dest_c.font = Font(name=src_c.font.name, size=src_c.font.size, bold=src_c.font.bold)
                    dest_c.border = copy(src_c.border)
                    dest_c.alignment = copy(src_c.alignment)
                 with contextlib.suppress(Exception): ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

            qty = float(item.get("quantity") or item.get("dsp_qty") or 0)
            rate = float(item.get("rate") or 0)
            taxable = float(item.get("taxable_value", qty * rate))
            cgst = float(item.get("cgst_amount", 0))
            sgst = float(item.get("sgst_amount", 0))
            line_total = float(item.get("total_amount", taxable + cgst + sgst))

            t_qty += qty
            t_taxable += taxable
            t_cgst += cgst
            t_sgst += sgst
            t_total += line_total

            set_val(f"A{r}", idx + 1, align="center")
            set_val(f"B{r}", item.get("material_description") or item.get("description"), align="left")
            set_val(f"I{r}", item.get("hsn_sac") or "")
            set_val(f"J{r}", item.get("material_code") or "", align="center")
            set_val(f"K{r}", item.get("no_of_packets", 1), align="center")
            set_val(f"L{r}", qty, align="center")
            set_val(f"M{r}", rate, align="center")
            set_val(f"N{r}", item.get("unit", "NOS"), align="center")
            set_val(f"O{r}", taxable, align="right")
            set_val(f"P{r}", cgst_rate, align="center")
            set_val(f"Q{r}", cgst, align="right")
            set_val(f"R{r}", sgst_rate, align="center")
            set_val(f"S{r}", sgst, align="right")
            set_val(f"T{r}", line_total, align="right")
        
        final_total_row = template_total_row + rows_to_insert
        set_val(f"H{final_total_row}", "Total", bold=True, align="right") 
        set_val(f"L{final_total_row}", t_qty)
        set_val(f"O{final_total_row}", t_taxable)
        set_val(f"Q{final_total_row}", t_cgst)
        set_val(f"S{final_total_row}", t_sgst)
        set_val(f"T{final_total_row}", t_total)

        set_val(f"A{final_total_row + 1}", f"Total Amount (In Words):- {amount_to_words(t_total)}", align="left")
        
        tax_sum_row = final_total_row + 3 
        set_val(f"O{tax_sum_row}", t_taxable)
        set_val(f"P{tax_sum_row}", cgst_rate)
        set_val(f"Q{tax_sum_row}", t_cgst)
        set_val(f"R{tax_sum_row}", sgst_rate)
        set_val(f"S{tax_sum_row}", t_sgst)
        set_val(f"T{tax_sum_row}", t_cgst + t_sgst)

        set_val(f"A{final_total_row + 5}", f"SGST (in words) : {amount_to_words(t_sgst)}", align="left")
        set_val(f"A{final_total_row + 6}", f"CGST (in words) : {amount_to_words(t_cgst)}", align="left")

        wb.save(self.output)
        self.output.seek(0)
        
        return self._save_or_stream(self.output, filename, save_path)